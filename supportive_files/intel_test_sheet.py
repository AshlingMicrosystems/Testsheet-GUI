import openpyxl
import sys
import configparser
import os
import shutil
import matplotlib.pyplot as plt
import numpy as np
import webbrowser
import time


#make directory to store chart images
dir_name = "static/images/"+sys.argv[1]
loc_wrong=sys.argv[2]
loc=loc_wrong.replace("*"," ")

if os.path.exists(dir_name):
    shutil.rmtree(dir_name)
os.makedirs(dir_name) 

# riscv.ini will be the deafult file if there is no commandline parameter
#if len(sys.argv)<=1:
IniFileName="Test_Sheet.ini"
#else:
#    IniFileName = sys.argv[1];

# Check commandline parameter file path/file are valid or not
# Print the error msg and exit the script if it is not valid
try:
    f = open(IniFileName)
    # Do something with the file
except IOError:
    print("Ini file does not exists/accessible in the following path:"+IniFileName)
    exit(0)

# reag the configurations from the ini file
read_config = configparser.ConfigParser()
read_config.read(IniFileName)

#test case type functionality, such as Base functionality, abnormal, stress etc...
TestCaseType=read_config.get("IntelTestSheet-parameters", "Base_fun_total").strip()
rowid = int(TestCaseType[1:])
ColumnChar = TestCaseType[0]
colid=ord(ColumnChar)-64

# Test selection reading
TestCaseFunction1       = read_config.get("IntelTestCase-parameters", "sheet1").strip()
TestCaseFunction2       = read_config.get("IntelTestCase-parameters", "sheet2").strip()
TestCaseFunction3	    = read_config.get("IntelTestCase-parameters", "sheet3").strip()
TestCaseFunction4	    = read_config.get("IntelTestCase-parameters", "sheet4").strip()
TestCaseFunction5       = read_config.get("IntelTestCase-parameters", "sheet5").strip()
TestCaseFunction6       = read_config.get("IntelTestCase-parameters", "sheet6").strip()
TestCaseFunction7	    = read_config.get("IntelTestCase-parameters", "sheet7").strip()
TestCaseFunction8       = read_config.get("IntelTestCase-parameters", "sheet8").strip()
TestCaseFunction9       = read_config.get("IntelTestCase-parameters", "sheet9").strip()

value=[]
#loc ="RiscFreeTestCases_Intel_Win.xlsm"
print(loc)
#exit()
wb = openpyxl.load_workbook(loc,data_only=True)
sheet = wb["Summary-Sheet"]

colors = ['green','red','grey']
explode = (0.1, 0.1, 0.1)             
def TestCaseType(rowid,colid):
    print("Total = ",sheet.cell(row=(rowid),column=(colid)).internal_value)
    print("Pass = ",sheet.cell(row=(rowid),column=(colid+1)).internal_value)
    print("Fail = ",sheet.cell(row=(rowid),column=(colid+2)).internal_value)
    print("Untested = ",sheet.cell(row=(rowid),column=(colid+3)).internal_value)
    value.append(sheet.cell(row=(rowid),column=(colid+1)).internal_value)
    value.append(sheet.cell(row=(rowid),column=(colid+2)).internal_value)
    value.append(sheet.cell(row=(rowid),column=(colid+3)).internal_value)


def TestType(rowid,colid,functionality):
    print("\n"+functionality)
    if (functionality=="Base Functionality"):
        rowid=rowid
    if (functionality=="Abnormal"):
        rowid=rowid+1
    if (functionality=="Performance"):
        rowid=rowid+2
    if (functionality=="Stress"):
        rowid=rowid+3
    if (functionality=="Installation"):
        rowid=rowid+4
    if (functionality=="ErrorHandling"):
        rowid=rowid+5
    if (functionality=="TotalTestCases"):
        rowid=rowid+6
    value.clear()
    TestCaseType(rowid,colid)
    print("\n")
    mylabels = ["Pass", "Fail", "Untested"]
    y = np.array(value[0:3])
    def make_autopct(values):
        def my_autopct(pct):
            total = sum(values)
            val = int(round(pct*total/100.0))
            return '{p:.2f}%  ({v:d})'.format(p=pct,v=val)
        return my_autopct
    plt.pie(y, autopct=make_autopct(y), colors=colors, explode=explode)
    plt.legend(title = functionality+":", labels = mylabels, bbox_to_anchor =(0.90, 1.15), ncol = 3)
    plt.savefig(dir_name+'/'+functionality+'.png')
    plt.close()

TestType(rowid,colid,"Base Functionality")
TestType(rowid,colid,"Abnormal")
TestType(rowid,colid,"Performance")
TestType(rowid,colid,"Stress")
TestType(rowid,colid,"Installation")
TestType(rowid,colid,"ErrorHandling")
TestType(rowid,colid,"TotalTestCases")

def TestSheetPlot(TestCaseName,TestCaseCell):
    rowid = int(TestCaseCell[1:])
    ColumnChar = TestCaseCell[0]
    colid=ord(ColumnChar)-64
    print("\n"+TestCaseName)
    value.clear()
    TestCaseType(rowid,colid)
    print("\n")
    mylabels = ["Pass", "Fail", "Untested"]
    y = np.array(value[0:3])
    def make_autopct(values):
        def my_autopct(pct):
            total = sum(values)
            val = int(round(pct*total/100.0))
            return '{p:.2f}%  ({v:d})'.format(p=pct,v=val)
        return my_autopct

    plt.pie(y, autopct=make_autopct(y), colors=colors, explode=explode)
    plt.legend(title = TestCaseName+":", labels = mylabels, bbox_to_anchor =(0.90, 1.15), ncol = 3)
    plt.savefig(dir_name+'/'+TestCaseName+'.png')
    plt.close()

if int(TestCaseFunction1)==1 : 
    value.clear()
    TestCaseName=read_config.get("IntelTestCase-parameters", "sheet1_name").strip()
    TestCaseCell=read_config.get("IntelTestCase-parameters", "sheet1_total").strip()
    TestSheetPlot(TestCaseName,TestCaseCell)

if int(TestCaseFunction2)==1 : 
    value.clear()
    TestCaseName=read_config.get("IntelTestCase-parameters", "sheet2_name").strip()
    TestCaseCell=read_config.get("IntelTestCase-parameters", "sheet2_total").strip()
    TestSheetPlot(TestCaseName,TestCaseCell)

if int(TestCaseFunction3)==1 : 
    value.clear()
    TestCaseName=read_config.get("IntelTestCase-parameters", "sheet3_name").strip()
    TestCaseCell=read_config.get("IntelTestCase-parameters", "sheet3_total").strip()
    TestSheetPlot(TestCaseName,TestCaseCell)

if int(TestCaseFunction4)==1 : 
    value.clear()
    TestCaseName=read_config.get("IntelTestCase-parameters", "sheet4_name").strip()
    TestCaseCell=read_config.get("IntelTestCase-parameters", "sheet4_total").strip()
    TestSheetPlot(TestCaseName,TestCaseCell)

if int(TestCaseFunction5)==1 : 
    value.clear()
    TestCaseName=read_config.get("IntelTestCase-parameters", "sheet5_name").strip()
    TestCaseCell=read_config.get("IntelTestCase-parameters", "sheet5_total").strip()
    TestSheetPlot(TestCaseName,TestCaseCell)

if int(TestCaseFunction6)==1 : 
    value.clear()
    TestCaseName=read_config.get("IntelTestCase-parameters", "sheet6_name").strip()
    TestCaseCell=read_config.get("IntelTestCase-parameters", "sheet6_total").strip()
    TestSheetPlot(TestCaseName,TestCaseCell)

if int(TestCaseFunction7)==1 : 
    value.clear()
    TestCaseName=read_config.get("IntelTestCase-parameters", "sheet7_name").strip()
    TestCaseCell=read_config.get("IntelTestCase-parameters", "sheet7_total").strip()
    TestSheetPlot(TestCaseName,TestCaseCell)

if int(TestCaseFunction8)==1 : 
    value.clear()
    TestCaseName=read_config.get("IntelTestCase-parameters", "sheet8_name").strip()
    TestCaseCell=read_config.get("IntelTestCase-parameters", "sheet8_total").strip()
    TestSheetPlot(TestCaseName,TestCaseCell)

if int(TestCaseFunction9)==1 : 
    value.clear()
    TestCaseName=read_config.get("IntelTestCase-parameters", "sheet9_name").strip()
    TestCaseCell=read_config.get("IntelTestCase-parameters", "sheet9_total").strip()
    TestSheetPlot(TestCaseName,TestCaseCell)

#Opening the html using browser
#if (sys.argv[1]=="chart_win"):
    #webbrowser.open_new_tab('Test_Status_update_win.html')
#    webbrowser.open("file://" + os.path.realpath("supportive_files/html/Test_Status_update_win.html"))
#if (sys.argv[1]=="chart_lin"):
    #webbrowser.open_new_tab('Test_Status_update_lin.html')
#    webbrowser.open("file://" + os.path.realpath("supportive_files/html/Test_Status_update_lin.html"))